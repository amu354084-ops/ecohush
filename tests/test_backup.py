import sqlite3
import hashlib
import json
from datetime import datetime
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

from app.services import backup


def test_create_local_backup_creates_timestamped_zip(tmp_path, monkeypatch):
    database = tmp_path / "erp_local.db"
    database.write_bytes(b"sqlite test database")
    monkeypatch.setenv("BACKUP_DIR", str(tmp_path / "backups"))
    monkeypatch.setattr(backup, "database_path", lambda: database)

    result = backup.create_local_backup(notify_telegram=False)

    archive = Path(result["path"])
    assert archive.exists()
    assert archive.name.startswith("backup_")
    assert archive.name.endswith(".zip")
    with ZipFile(archive) as zip_file:
        assert zip_file.read("erp_local.db") == b"sqlite test database"
        assert "описание_резервной_копии.json" in zip_file.namelist()
        manifest = json.loads(zip_file.read("описание_резервной_копии.json"))
        assert manifest["sha256_базы"] == hashlib.sha256(b"sqlite test database").hexdigest()


def test_create_local_backup_exports_formatted_excel(tmp_path, monkeypatch):
    database = tmp_path / "erp_local.db"
    with sqlite3.connect(database) as connection:
        connection.executescript(
            """
            CREATE TABLE items (id INTEGER PRIMARY KEY, code TEXT, name TEXT, type TEXT, unit TEXT, min_stock INTEGER, price NUMERIC);
            CREATE TABLE warehouses (id INTEGER PRIMARY KEY, name TEXT);
            CREATE TABLE batches (id INTEGER PRIMARY KEY, item_id INTEGER, warehouse_id INTEGER, purchase_cost NUMERIC, initial_qty NUMERIC, remaining_qty NUMERIC, created_at TEXT);
            CREATE TABLE stock_transactions (id INTEGER PRIMARY KEY, batch_id INTEGER, type TEXT, qty NUMERIC, comment TEXT, timestamp TEXT);
            CREATE TABLE users (id INTEGER PRIMARY KEY, username TEXT, full_name TEXT, role TEXT, can_change_status INTEGER, is_active INTEGER);
            CREATE TABLE counterparties (id INTEGER PRIMARY KEY, name TEXT, phone TEXT, current_debt NUMERIC);
            CREATE TABLE orders (id INTEGER PRIMARY KEY, invoice_number TEXT, courier_id INTEGER, client_id INTEGER, status TEXT, payment_type TEXT, discount_amount NUMERIC, created_at TEXT, delivered_at TEXT);
            CREATE TABLE order_items (id INTEGER PRIMARY KEY, order_id INTEGER, item_id INTEGER, quantity NUMERIC, price NUMERIC, discount NUMERIC);
            CREATE TABLE cash_transactions (id INTEGER PRIMARY KEY, type TEXT, amount NUMERIC, payment_method TEXT, counterparty_id INTEGER, description TEXT, created_at TEXT);
            INSERT INTO items VALUES (1, 'A-1', 'Товар', 'FINAL', 'шт', 2, 125.50);
            INSERT INTO warehouses VALUES (1, 'Основной');
            INSERT INTO batches VALUES (1, 1, 1, 80, 10, 7, '2026-08-22 10:00:00');
            INSERT INTO stock_transactions VALUES (1, 1, 'INBOUND', 10, 'Поставка', '2026-08-22 10:00:00');
            INSERT INTO users VALUES (1, 'admin', 'Администратор', 'ADMIN', 1, 1);
            INSERT INTO counterparties VALUES (1, 'Клиент', '+70000000000', 300.25);
            INSERT INTO orders VALUES (1, 'INV-1', 1, 1, 'DELIVERED', 'DEBT', 5, '2026-08-22 11:00:00', NULL);
            INSERT INTO order_items VALUES (1, 1, 1, 2, 125.50, 0);
            INSERT INTO cash_transactions VALUES (1, 'INCOME', 100, 'CASH', 1, 'Оплата', '2026-08-22 12:00:00');
            """
        )
    monkeypatch.setenv("BACKUP_DIR", str(tmp_path / "backups"))
    monkeypatch.setattr(backup, "database_path", lambda: database)

    result = backup.create_local_backup(notify_telegram=False)

    excel_path = Path(result["excel_path"])
    assert excel_path.exists()
    assert excel_path.parent.name == f"excel_{datetime.now():%Y_%m_%d}"
    workbook = load_workbook(excel_path, data_only=False)
    assert {
        "Склад",
        "Движение склада",
        "Заказы",
        "Позиции заказов",
        "Касса",
        "Операции кассы",
        "Должники",
        "Пользователи и роли",
    } == set(workbook.sheetnames)
    assert workbook["Склад"]["C2"].value == "Товар"
    assert workbook["Позиции заказов"]["E2"].value == "Товар"
    assert workbook["Касса"]["A3"].value == 100
    assert workbook["Касса"].auto_filter.ref == "A2:A3"
    assert workbook["Должники"]["D2"].number_format == '#,##0.00;[Red]-#,##0.00'
