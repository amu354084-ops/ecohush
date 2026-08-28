"""Convert an xlsx file to csv (utf-8-sig) and print first N lines.
Usage: python tools/convert_and_preview.py exports/invoice_shipment_8.xlsx 10
"""
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception:
    print("openpyxl not installed. Install with: pip install openpyxl")
    raise

p = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("exports/invoice_shipment_8.xlsx")
lines = int(sys.argv[2]) if len(sys.argv) > 2 else 10
if not p.exists():
    print("File not found:", p)
    sys.exit(2)

wb = load_workbook(p, read_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
if not rows:
    print("Empty sheet")
    sys.exit(0)

csv_path = p.with_suffix('.csv')
import csv
with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
    w = csv.writer(f)
    for r in rows:
        w.writerow(["" if v is None else v for v in r])

print("Saved CSV:", csv_path)
print("--- Preview ---")
for i, r in enumerate(rows[:lines], start=1):
    cleaned = ["" if v is None else v for v in r]
    print(i, ", ".join([str(v) for v in cleaned]))
