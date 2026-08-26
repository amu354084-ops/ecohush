"""Inspect first row (headers) of the exported Excel file.
Usage: python tools/inspect_export.py exports/invoice_shipment_8.xlsx
"""
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception:
    print("openpyxl not installed. Install with: pip install openpyxl")
    raise

p = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("exports/invoice_shipment_8.xlsx")
if not p.exists():
    print("File not found:", p)
    sys.exit(2)

wb = load_workbook(p, read_only=True)
ws = wb[wb.sheetnames[0]]
first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
print("Headers:")
for i, h in enumerate(first_row, start=1):
    print(i, repr(h))
