from __future__ import annotations

import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


EXCEL_FILENAME = "ERP_Полный_Архив.xlsx"

FIELD_LABELS = {
    "id": "Идентификатор",
    "sale_id": "Продажа",
    "shipment_id": "Отгрузка",
    "item_id": "Товар",
    "batch_id": "Партия",
    "qty": "Количество",
    "unit_price": "Цена за единицу",
    "cost_price": "Себестоимость",
    "discount_percent": "Скидка, %",
    "created_at": "Дата создания",
    "status": "Статус",
    "note": "Примечание",
    "code": "Код",
    "name": "Товар",
    "type": "Тип",
    "unit": "Единица измерения",
    "min_stock": "Минимальный остаток",
    "price": "Цена",
    "purchase_cost": "Закупочная цена",
    "initial_qty": "Начальное количество",
    "remaining_qty": "Остаток",
    "warehouse_id": "Склад",
    "warehouse_name": "Склад",
    "item_code": "Код товара",
    "item_name": "Товар",
    "comment": "Комментарий",
    "timestamp": "Дата и время",
    "invoice_number": "Номер накладной",
    "payment_type": "Тип оплаты",
    "discount_amount": "Сумма скидки",
    "courier": "Курьер",
    "client_name": "Клиент",
    "client_phone": "Телефон клиента",
    "quantity": "Количество",
    "discount": "Скидка",
    "line_total": "Сумма позиции",
}


def _read_query(connection: sqlite3.Connection, query: str) -> pd.DataFrame:
    dataframe = pd.read_sql_query(query, connection)
    for column in dataframe.columns:
        if "date" in str(column).lower() or "time" in str(column).lower() or column in {"created_at", "timestamp"}:
            dataframe[column] = pd.to_datetime(dataframe[column], errors="coerce")
            if isinstance(dataframe[column].dtype, pd.DatetimeTZDtype):
                dataframe[column] = dataframe[column].dt.tz_localize(None)
            elif dataframe[column].dtype == object:
                dataframe[column] = dataframe[column].map(
                    lambda value: value.replace(tzinfo=None)
                    if hasattr(value, "tzinfo") and value.tzinfo is not None
                    else value
                )
    return dataframe


def _format_sheet(worksheet, header_row: int = 1) -> None:
    worksheet.freeze_panes = f"A{header_row + 1}"
    last_column = get_column_letter(worksheet.max_column)
    worksheet.auto_filter.ref = f"A{header_row}:{last_column}{worksheet.max_row}"
    worksheet.sheet_view.showGridLines = False

    for cell in worksheet[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 42)

    for row in worksheet.iter_rows(min_row=header_row + 1):
        for cell in row:
            header = str(worksheet.cell(header_row, cell.column).value or "").lower()
            if "date" in header or "time" in header:
                cell.number_format = "yyyy-mm-dd hh:mm"
            elif any(word in header for word in ("amount", "price", "cost", "debt", "balance", "total")):
                cell.number_format = "#,##0.00;[Red]-#,##0.00"


def _write_dataframe(writer: pd.ExcelWriter, dataframe: pd.DataFrame, sheet_name: str) -> None:
    _russian_columns(dataframe).to_excel(writer, sheet_name=sheet_name, index=False)


def _russian_columns(dataframe: pd.DataFrame) -> pd.DataFrame:
    renamed = {
        column: FIELD_LABELS.get(column, column.replace("_", " ").capitalize())
        for column in dataframe.columns
    }
    return dataframe.rename(columns=renamed)


def _read_additional_tables(connection: sqlite3.Connection) -> dict[str, pd.DataFrame]:
    sheet_names = {
        "bom_headers": "BOM заголовки",
        "bom_items": "BOM состав",
        "sales": "Продажи",
        "sale_items": "Позиции продаж",
        "overhead_expenses": "Накладные расходы",
        "shipments": "Отгрузки",
        "shipment_items": "Позиции отгрузок",
    }
    tables = connection.execute(
        "SELECT name FROM sqlite_master WHERE type = 'table' AND name NOT LIKE 'sqlite_%'"
    ).fetchall()
    exported_tables = {
        "items",
        "batches",
        "stock_transactions",
        "orders",
        "order_items",
        "cash_transactions",
        "counterparties",
        "users",
    }
    return {
        sheet_names[table_name]: _russian_columns(_read_query(connection, f'SELECT * FROM "{table_name}"'))
        for (table_name,) in tables
        if table_name not in exported_tables and table_name in sheet_names
    }


def export_database_to_excel(database_path: Path, output_directory: Path) -> Path:
    """Export the operational database into a readable, formatted workbook."""
    output_directory.mkdir(parents=True, exist_ok=True)
    output_path = output_directory / EXCEL_FILENAME

    with sqlite3.connect(database_path) as connection:
        warehouse = _read_query(
            connection,
            """
            SELECT i.id AS item_id, i.code, i.name, i.type, i.unit, i.min_stock, i.price,
                   b.id AS batch_id, b.purchase_cost, b.initial_qty, b.remaining_qty,
                   b.created_at AS batch_created_at, w.id AS warehouse_id, w.name AS warehouse_name
            FROM items i
            LEFT JOIN batches b ON b.item_id = i.id
            LEFT JOIN warehouses w ON w.id = b.warehouse_id
            ORDER BY i.name, w.name, b.id
            """,
        )
        stock_transactions = _read_query(
            connection,
            """
            SELECT st.id, st.batch_id, i.code AS item_code, i.name AS item_name,
                   w.name AS warehouse_name, st.type, st.qty, st.comment, st.timestamp
            FROM stock_transactions st
            JOIN batches b ON b.id = st.batch_id
            JOIN items i ON i.id = b.item_id
            JOIN warehouses w ON w.id = b.warehouse_id
            ORDER BY st.timestamp, st.id
            """,
        )
        orders = _read_query(
            connection,
            """
            SELECT o.id, o.invoice_number, o.status, o.payment_type, o.discount_amount,
                   o.created_at, o.delivered_at, u.username AS courier,
                   c.name AS client_name, c.phone AS client_phone,
                   COALESCE(SUM(oi.quantity * oi.price - oi.discount), 0) - o.discount_amount AS total_amount
            FROM orders o
            LEFT JOIN order_items oi ON oi.order_id = o.id
            LEFT JOIN users u ON u.id = o.courier_id
            LEFT JOIN counterparties c ON c.id = o.client_id
            GROUP BY o.id
            ORDER BY o.created_at, o.id
            """,
        )
        order_items = _read_query(
            connection,
            """
            SELECT oi.id, oi.order_id, o.invoice_number, i.code AS item_code, i.name AS item_name,
                   i.unit, oi.quantity, oi.price, oi.discount,
                   (oi.quantity * oi.price - oi.discount) AS line_total
            FROM order_items oi
            JOIN orders o ON o.id = oi.order_id
            JOIN items i ON i.id = oi.item_id
            ORDER BY oi.order_id, oi.id
            """,
        )
        cash = _read_query(
            connection,
            """
            SELECT ct.id, ct.type, ct.amount, ct.payment_method, ct.created_at,
                   c.name AS counterparty_name, ct.description
            FROM cash_transactions ct
            LEFT JOIN counterparties c ON c.id = ct.counterparty_id
            ORDER BY ct.created_at, ct.id
            """,
        )
        cash_balance = _read_query(
            connection,
            """
            SELECT COALESCE(SUM(CASE WHEN type = 'INCOME' THEN amount ELSE -amount END), 0) AS balance
            FROM cash_transactions
            """,
        )
        debtors = _read_query(
            connection,
            """
            SELECT id, name, phone, current_debt
            FROM counterparties
            WHERE current_debt <> 0
            ORDER BY current_debt DESC, name
            """,
        )
        users = _read_query(
            connection,
            """
            SELECT id, username, full_name, role, can_change_status, is_active
            FROM users
            ORDER BY username
            """,
        )
        additional_tables = _read_additional_tables(connection)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _write_dataframe(writer, warehouse, "Склад")
        _write_dataframe(writer, stock_transactions, "Движение склада")
        _write_dataframe(writer, orders, "Заказы")
        _write_dataframe(writer, order_items, "Позиции заказов")
        cash_balance.rename(columns={"balance": "Текущий баланс"}).to_excel(
            writer, sheet_name="Касса", index=False, startrow=1
        )
        _write_dataframe(writer, cash, "Операции кассы")
        _write_dataframe(writer, debtors, "Должники")
        _write_dataframe(writer, users, "Пользователи и роли")
        for sheet_name, dataframe in additional_tables.items():
            _write_dataframe(writer, dataframe, sheet_name)

    workbook = load_workbook(output_path)
    for worksheet in workbook.worksheets:
        _format_sheet(worksheet, header_row=2 if worksheet.title == "Касса" else 1)
    workbook["Касса"]["A1"] = "Текущий баланс"
    workbook["Касса"]["A1"].font = Font(bold=True, color="FFFFFF")
    workbook["Касса"]["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    workbook.save(output_path)
    return output_path
